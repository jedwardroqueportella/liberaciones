import os
import re
from flask import Flask, render_template, request, send_file
import mysql.connector
from mysql.connector import Error
import openpyxl
from io import BytesIO

DB_CONFIG = {
    "host": "localhost",
    "port": 3306,
    "user": "admin",
    "password": "O-7Qou#%5Np2",
    "database": "Inventario_Admin_IP_TheFinal",
    "auth_plugin": "mysql_native_password",
}

RCSR_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "rCSR")

try:
    with open(RCSR_TEMPLATE_PATH, "r", encoding="utf-8") as f:
        RCSR_TEMPLATE = f.read()
except Exception:
    RCSR_TEMPLATE = ""

TOPOLOGIES = {
    "1": "AGG - rCSR",
    "2": "AGG - preagregador - rCSR",
    "3": "AGG - MW - rCSR",
}

REGION_ZONE = {
    "AMAZONAS": "NORTE",
    "ANCASH": "NORTE",
    "APURIMAC": "SUR",
    "AREQUIPA": "SUR",
    "AYACUCHO": "CENTRO",
    "CAJAMARCA": "NORTE",
    "CUSCO": "SUR",
    "HUANCAVELICA": "CENTRO",
    "HUANUCO": "CENTRO",
    "ICA": "SUR",
    "JUNIN": "CENTRO",
    "LA LIBERTAD": "NORTE",
    "LAMBAYEQUE": "NORTE",
    "LIMA": "LIMA",
    "LORETO": "CENTRO",
    "MADRE DE DIOS": "SUR",
    "MOQUEGUA": "SUR",
    "PASCO": "CENTRO",
    "PIURA": "NORTE",
    "PUNO": "SUR",
    "SAN MARTIN": "NORTE",
    "TACNA": "SUR",
    "TUMBES": "NORTE",
    "UCAYALI": "CENTRO",
}

ZONE_COMMUNITY = {
    "NORTE": "65002",
    "CENTRO": "65004",
    "LIMA": "65001",
    "SUR": "65003",
}

app = Flask(__name__)


def get_connection():
    return mysql.connector.connect(**DB_CONFIG)


def get_zone(region):
    if not region:
        return "N/A"
    return REGION_ZONE.get(region.strip().upper(), "N/A")


def normalize_router(row, source):
    if source == "cellsiterouter":
        return {
            "type": "rCSR",
            "hostname": row["hostname"],
            "loopback_0": row["loopback_0"],
            "loopback_1": row["loopback_1"],
            "ipv6": row.get("loopback0_ipv6") or "---",
            "region": row["region"],
            "zone": get_zone(row["region"]),
        }

    return {
        "type": "AGG",
        "hostname": row["equipo_agregador"],
        "loopback_0": row["ip_sistema"],
        "loopback_1": row["ip_gestion"],
        "ipv6": row.get("loopback0_ipv6") or "---",
        "region": row["region"],
        "zone": get_zone(row["region"]),
    }


def get_agregator_info(agg_hostname):
    if not agg_hostname:
        return None, "Debe indicar el hostname del AGG."

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT a.equipo_agregador, a.ip_sistema, a.ip_gestion, a.region, ipv.loopback0_ipv6 "
            "FROM agregattion a "
            "LEFT JOIN Agg_ipv6 ipv ON LOWER(ipv.ipran) = LOWER(a.equipo_agregador) "
            "WHERE LOWER(a.equipo_agregador) = %s",
            (agg_hostname.lower(),),
        )
        row = cursor.fetchone()
        if not row:
            return None, f"No se encontró el AGG '{agg_hostname}' en la base de datos."

        zone = get_zone(row["region"])
        community = ZONE_COMMUNITY.get(zone, "N/A")
        return {
            "agg_hostname": row["equipo_agregador"],
            "agg_loopback0": row["ip_sistema"],
            "agg_loopback1": row["ip_gestion"],
            "agg_loopback0_ipv6": row.get("loopback0_ipv6") or "",
            "agg_region": row["region"],
            "agg_zone": zone,
            "agg_community": community,
        }, None
    except Error as exc:
        return None, f"Error de conexión a la base de datos: {exc}"
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def get_preagg_info(preagg_hostname):
    if not preagg_hostname:
        return None, "Debe indicar el hostname del preagregador."

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT hostname, loopback_0 FROM cellsiterouter WHERE LOWER(hostname) = %s",
            (preagg_hostname.lower(),),
        )
        row = cursor.fetchone()
        if not row:
            return None, f"No se encontró el preagregador '{preagg_hostname}' en la base de datos."

        return {
            "preagg_hostname": row["hostname"],
            "preagg_loopback0": row["loopback_0"],
        }, None
    except Error as exc:
        return None, f"Error de conexión a la base de datos: {exc}"
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def generate_rcsr_script(data):
    values = {
        "rcsr_hostname": ((data.get("rcsr_hostname") or data.get("csr_hostname")) or "").strip(),
        "rcsr_port": ((data.get("rcsr_port") or data.get("csr_port")) or "").strip(),
        "csr_loopback0": (data.get("csr_loopback0") or "").strip(),
        "csr_loopback1": (data.get("csr_loopback1") or "").strip(),
        "wan_ip": (data.get("wan_ip") or "").strip(),
        "ospf_area": (data.get("ospf_area") or "").strip(),
        "agg_hostname": (data.get("agg_hostname") or "").strip(),
        "agg_port": (data.get("agg_port") or "").strip(),
        "agg_loopback0": (data.get("agg_loopback0") or "").strip(),
        "agg_zone": (data.get("agg_zone") or "").strip(),
        "agg_community": (data.get("agg_community") or "").strip(),
        "cli": (data.get("cli") or "").strip(),
        "cidpext": (data.get("cidpext") or "").strip(),
        "ide": (data.get("ide") or "").strip(),
        "mpls_password": (data.get("mpls_password") or "B&k9DO%d4").strip() or "B&k9DO%d4",
    }

    if not values["agg_community"] and values["agg_zone"]:
        values["agg_community"] = ZONE_COMMUNITY.get(values["agg_zone"].upper(), "")

    required = [
        "rcsr_hostname",
        "rcsr_port",
        "csr_loopback0",
        "csr_loopback1",
        "wan_ip",
        "ospf_area",
        "agg_hostname",
        "agg_port",
        "agg_loopback0",
        "agg_community",
        "cli",
        "cidpext",
        "ide",
    ]

    missing = [name for name in required if not values[name]]
    if missing:
        missing_labels = ", ".join(missing)
        return None, f"Faltan campos requeridos en rCSR: {missing_labels}."

    if not RCSR_TEMPLATE:
        return None, "Plantilla rCSR no encontrada."

    replacements = {
        "rCSRElSoldeNaranjal2": values["rcsr_hostname"],
        "10.253.0.220": values["csr_loopback1"],
        "10.253.32.220": values["csr_loopback0"],
        "10.253.156.52": values["wan_ip"],
        "5.14.1.51": values["ospf_area"],
        "770941": values["ide"],
        "25940352": values["cidpext"],
        "LI1264": values["cli"],
        "ASG-LIM-LosOlivos2": values["agg_hostname"],
        "INTERFACE_DESC_HOSTNAME": values["agg_hostname"],
        "INTERFACE_DESC_PORT": values["agg_port"],
        "LDP_PEER_LOOPBACK": values["agg_loopback0"],
        "10.140.132.58": values["agg_loopback0"],
        "B&k9DO%d4": values["mpls_password"],
        "65001": values["agg_community"],
    }

    script = RCSR_TEMPLATE
    for sample, actual in replacements.items():
        script = script.replace(sample, actual)

    return script, None


def generate_rcsr_script_agg_bgp(data):
    # Para topología 2, BGP apunta al ASG (peer loopback, no necesita puerto físico)
    values = {
        "rcsr_hostname": ((data.get("rcsr_hostname") or data.get("csr_hostname")) or "").strip(),
        "csr_loopback0": (data.get("csr_loopback0") or "").strip(),
        "csr_loopback1": (data.get("csr_loopback1") or "").strip(),
        "wan_ip": (data.get("wan_ip") or "").strip(),
        "ospf_area": (data.get("ospf_area") or "").strip(),
        "agg_hostname": (data.get("agg_hostname") or "").strip(),
        "agg_loopback0": (data.get("agg_loopback0") or "").strip(),
        "preagg_hostname": (data.get("preagg_hostname") or "").strip(),
        "preagg_port": (data.get("preagg_port") or "").strip(),
        "preagg_loopback0": (data.get("preagg_loopback0") or "").strip(),
        "agg_zone": (data.get("agg_zone") or "").strip(),
        "agg_community": (data.get("agg_community") or "").strip(),
        "cli": (data.get("cli") or "").strip(),
        "cidpext": (data.get("cidpext") or "").strip(),
        "ide": (data.get("ide") or "").strip(),
        "mpls_password": (data.get("mpls_password") or "B&k9DO%d4").strip() or "B&k9DO%d4",
    }

    if not values["agg_community"] and values["agg_zone"]:
        values["agg_community"] = ZONE_COMMUNITY.get(values["agg_zone"].upper(), "")

    required = [
        "rcsr_hostname",
        "csr_loopback0",
        "csr_loopback1",
        "wan_ip",
        "ospf_area",
        "agg_hostname",
        "agg_loopback0",
        "preagg_hostname",
        "preagg_port",
        "preagg_loopback0",
        "agg_community",
        "cli",
        "cidpext",
        "ide",
    ]

    missing = [name for name in required if not values[name]]
    if missing:
        missing_labels = ", ".join(missing)
        return None, f"Faltan campos requeridos en rCSR: {missing_labels}."

    if not RCSR_TEMPLATE:
        return None, "Plantilla rCSR no encontrada."

    replacements = {
        "rCSRElSoldeNaranjal2": values["rcsr_hostname"],
        "10.253.0.220": values["csr_loopback1"],
        "10.253.32.220": values["csr_loopback0"],
        "10.253.156.52": values["wan_ip"],
        "5.14.1.51": values["ospf_area"],
        "770941": values["ide"],
        "25940352": values["cidpext"],
        "LI1264": values["cli"],
        "ASG-LIM-LosOlivos2": values["agg_hostname"],
        "INTERFACE_DESC_HOSTNAME": values["preagg_hostname"],
        "INTERFACE_DESC_PORT": values["preagg_port"],
        "LDP_PEER_LOOPBACK": values["preagg_loopback0"],
        "10.140.132.58": values["agg_loopback0"],  # BGP peer remains AGG loopback
        "B&k9DO%d4": values["mpls_password"],
        "65001": values["agg_community"],
    }

    script = RCSR_TEMPLATE
    for sample, actual in replacements.items():
        script = script.replace(sample, actual)

    return script, None


def generate_preagg_script(data):
    values = {
        "preagg_hostname": (data.get("preagg_hostname") or "").strip(),
        "preagg_port": (data.get("preagg_port") or "").strip(),
        "preagg_loopback0": (data.get("preagg_loopback0") or "").strip(),
        "rcsr_hostname": ((data.get("rcsr_hostname") or data.get("csr_hostname")) or "").strip(),
        "rcsr_port": ((data.get("rcsr_port") or data.get("csr_port")) or "").strip(),
        "wan_ip": (data.get("wan_ip") or "").strip(),
        "csr_loopback0": (data.get("csr_loopback0") or "").strip(),
        "ospf_area": (data.get("ospf_area") or "").strip(),
        "cli": (data.get("cli") or "").strip(),
        "cidpext": (data.get("cidpext") or "").strip(),
        "ide": (data.get("ide") or "").strip(),
        "mpls_password": (data.get("mpls_password") or "B&k9DO%d4").strip() or "B&k9DO%d4",
    }

    required = [
        "preagg_hostname",
        "preagg_port",
        "preagg_loopback0",
        "rcsr_hostname",
        "rcsr_port",
        "wan_ip",
        "csr_loopback0",
        "ospf_area",
        "cli",
        "cidpext",
        "ide",
    ]

    missing = [name for name in required if not values[name]]
    if missing:
        missing_labels = ", ".join(missing)
        return None, f"Faltan campos requeridos en preagregador: {missing_labels}."

    script = f"""{values['preagg_hostname']}
system-view
interface GigabitEthernet {values['preagg_port']}
carrier up-hold-time 5000
mtu 9180
description IDE {values['ide']} CID {values['cidpext']} {values['cli']} CLARO PERU {values['rcsr_hostname']} {values['rcsr_port']}
undo shutdown
set flow-stat interval 30
control-flap
trust upstream transporte_IPRAN_diffserv
undo icmp name net-unreachable send
undo icmp name host-unreachable receive
undo icmp name protocol-unreachable receive
undo icmp name port-unreachable send
ospf cost 1000
ospf network-type p2p
ospf ldp-sync
ospf timer ldp-sync hold-max-cost infinite
ospf mtu-enable
mpls
mpls te
mpls rsvp-te
mpls rsvp-te hello
mpls rsvp-te srefresh compatible
mpls ldp
undo dcn
y
mpls ldp timer hello-send 3
mpls ldp timer keepalive-hold 45
y
mpls ldp timer keepalive-send 3
port-queue be wfq weight 5 port-wred best_effort outbound
port-queue af1 wfq weight 15 port-wred low-priority outbound
port-queue af2 wfq weight 40 port-wred high_priority outbound
port-queue af3 wfq weight 5 outbound
port-queue af4 wfq weight 5 outbound
port-queue ef pq shaping shaping-percentage 20 port-wred voice outbound
port-queue cs6 pq shaping shaping-percentage 10 port-wred signaling outbound
port-queue cs7 pq shaping shaping-percentage 1 port-wred signaling outbound
statistic enable
mpls poison-reverse enable
#
ospf 100
undo silent-interface GigabitEthernet {values['preagg_port']}
area {values['ospf_area']}
network {values['wan_ip']} 0.0.0.3 description to_{values['rcsr_hostname']}
#
mpls ldp
md5-password cipher {values['csr_loopback0']} {values['mpls_password']}
#
commit label Integracion_{values['rcsr_hostname']}
#
q
save"""

    return script, None


def generate_agg_with_rcsr_peer(data):
    values = {
        "agg_hostname": (data.get("agg_hostname") or "").strip(),
        "csr_loopback0": (data.get("csr_loopback0") or "").strip(),
        "rcsr_hostname": ((data.get("rcsr_hostname") or data.get("csr_hostname")) or "").strip(),
    }

    required = ["agg_hostname", "csr_loopback0", "rcsr_hostname"]
    missing = [name for name in required if not values[name]]
    if missing:
        missing_labels = ", ".join(missing)
        return None, f"Faltan campos requeridos en AGG: {missing_labels}."

    script = f"""{values['agg_hostname']}

#
bgp 12252
peer {values['csr_loopback0']} as-number 12252
peer {values['csr_loopback0']} group to_IPRAN_CSR
peer {values['csr_loopback0']} description to_{values['rcsr_hostname']}

ipv4-family unicast
peer {values['csr_loopback0']} enable
 y
peer {values['csr_loopback0']} group to_IPRAN_CSR

ipv6-family unicast
peer {values['csr_loopback0']} enable
 y
peer {values['csr_loopback0']} group to_IPRAN_CSR

ipv4-family vpnv4
peer {values['csr_loopback0']} enable
 y
peer {values['csr_loopback0']} group to_IPRAN_CSR
#
commit label Integracion_{values['rcsr_hostname']}
#
q
save"""

    return script, None


def create_excel(sheets):
    wb = openpyxl.Workbook()
    for i, (name, content) in enumerate(sheets):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        lines = content.split('\n')
        for row, line in enumerate(lines, 1):
            ws.cell(row=row, column=1).value = line
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def generate_agg_script(data):
    values = {
        "agg_hostname": (data.get("agg_hostname") or "").strip(),
        "agg_port": (data.get("agg_port") or "").strip(),
        "csr_hostname": ((data.get("csr_hostname") or data.get("rcsr_hostname")) or "").strip(),
        "csr_port": ((data.get("csr_port") or data.get("rcsr_port")) or "").strip(),
        "wan_ip": (data.get("wan_ip") or "").strip(),
        "csr_loopback0": (data.get("csr_loopback0") or "").strip(),
        "csr_loopback1": (data.get("csr_loopback1") or "").strip(),
        "ospf_area": (data.get("ospf_area") or "").strip(),
        "cli": (data.get("cli") or "").strip(),
        "cidpext": (data.get("cidpext") or "").strip(),
        "ide": (data.get("ide") or "").strip(),
        "mpls_password": (data.get("mpls_password") or "B&k9DO%d4").strip() or "B&k9DO%d4",
    }

    required = [
        "agg_hostname",
        "agg_port",
        "csr_hostname",
        "csr_port",
        "wan_ip",
        "csr_loopback0",
        "csr_loopback1",
        "ospf_area",
        "cli",
        "cidpext",
        "ide",
    ]

    missing = [name for name in required if not values[name]]
    if missing:
        missing_labels = ", ".join(missing)
        return None, f"Faltan campos requeridos: {missing_labels}."

    script = f"""{values['agg_hostname']}
system-view
interface GigabitEthernet {values['agg_port']}
carrier up-hold-time 5000
mtu 9180
description CID {values['cidpext']} >>{values['csr_hostname']} GI{values['csr_port']}
undo shutdown
set flow-stat interval 30
control-flap
trust upstream transporte_IPRAN_diffserv
undo icmp name net-unreachable send
undo icmp name host-unreachable receive
undo icmp name protocol-unreachable receive
undo icmp name port-unreachable send
ospf cost 1000
ospf network-type p2p
ospf ldp-sync
ospf timer ldp-sync hold-max-cost infinite
ospf mtu-enable
mpls
mpls te
mpls rsvp-te
mpls rsvp-te hello
mpls rsvp-te srefresh compatible
mpls ldp
undo dcn
y
mpls ldp timer hello-send 3
mpls ldp timer keepalive-hold 45
y
mpls ldp timer keepalive-send 3
port-queue be wfq weight 5 port-wred best_effort outbound
port-queue af1 wfq weight 15 port-wred low-priority outbound
port-queue af2 wfq weight 40 port-wred high_priority outbound
port-queue af3 wfq weight 5 outbound
port-queue af4 wfq weight 5 outbound
port-queue ef pq shaping shaping-percentage 20 port-wred voice outbound
port-queue cs6 pq shaping shaping-percentage 10 port-wred signaling outbound
port-queue cs7 pq shaping shaping-percentage 1 port-wred signaling outbound
statistic enable
mpls poison-reverse enable
#
ospf 110
area {values['ospf_area']}
network {values['wan_ip']} 0.0.0.3 description to_{values['csr_hostname']}
#
mpls ldp
md5-password cipher {values['csr_loopback0']} {values['mpls_password']}
#
bgp 12252
peer {values['csr_loopback0']} as-number 12252
peer {values['csr_loopback0']} group to_IPRAN_CSR
peer {values['csr_loopback0']} description to_{values['csr_hostname']}

ipv4-family unicast
peer {values['csr_loopback0']} enable
 y
peer {values['csr_loopback0']} group to_IPRAN_CSR

ipv6-family unicast
peer {values['csr_loopback0']} enable
 y
peer {values['csr_loopback0']} group to_IPRAN_CSR

ipv4-family vpnv4
peer {values['csr_loopback0']} enable
 y
peer {values['csr_loopback0']} group to_IPRAN_CSR
#
commit label Integracion_{values['csr_hostname']}
#
q
save"""

    return script, None


def parse_bulk_input(text):
    if not text:
        return {}

    lines = [line.strip() for line in text.replace('\r', '').splitlines() if line.strip()]
    if not lines:
        return {}

    header_pattern = re.compile(
        r'^\s*(Código|Codigo|Site|KM|Interface|Area OSPF|Device ZTE|Type Device|Loopback 0|Loopback 1|Locator|WANv6|ISIS Area|IPV6)',
        re.I,
    )
    data_lines = [line for line in lines if not header_pattern.match(line)]
    if not data_lines:
        return {}

    def split_parts(line):
        parts = re.split(r'\t+|\s{2,}', line)
        if len(parts) < 8:
            parts = re.split(r'\s+', line)
        return [part.strip() for part in parts if part.strip()]

    def find_ipv4_addresses(line):
        return re.findall(r'\b\d{1,3}(?:\.\d{1,3}){3}\b', line)

    parsed = {}
    main_line = None
    ipv6_line = None
    for line in data_lines:
        if '::' in line or re.search(r'[A-Fa-f0-9]{1,4}:[A-Fa-f0-9:]+', line):
            ipv6_line = line
            continue
        if main_line is None:
            main_line = line

    if not main_line:
        return {}

    parts = split_parts(main_line)
    if len(parts) >= 1:
        parsed["cli"] = parts[0]
    if len(parts) >= 4:
        parsed["agg_hostname"] = parts[3]
    if len(parts) >= 5:
        parsed["agg_port"] = parts[4]
    if len(parts) >= 6:
        parsed["csr_hostname"] = parts[5]
        parsed["rcsr_hostname"] = parts[5]
    if len(parts) >= 7:
        parsed["csr_port"] = parts[6]
        parsed["rcsr_port"] = parts[6]
    if len(parts) >= 8:
        parsed["wan_ip"] = parts[7]
    if len(parts) >= 9:
        parsed["csr_loopback1"] = parts[8]
    if len(parts) >= 10:
        parsed["csr_loopback0"] = parts[9]
    if len(parts) >= 11:
        parsed["ospf_area"] = parts[10]

    if ipv6_line:
        ipv6_addrs = re.findall(r'[0-9A-Fa-f:]+::[0-9A-Fa-f:]*', ipv6_line)
        if len(ipv6_addrs) >= 1:
            parsed["csr_loopback0_ipv6"] = ipv6_addrs[0]
        if len(ipv6_addrs) >= 2:
            parsed["locator"] = ipv6_addrs[1]
        if len(ipv6_addrs) >= 3:
            parsed["wan_ipv6"] = ipv6_addrs[2]
        parsed["has_ipv6"] = True

    return parsed


def find_single_router(router_name):
    router_name = (router_name or "").strip()
    if not router_name:
        return None

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        query_value = router_name.lower()

        cursor.execute(
            "SELECT hostname, loopback_0, loopback_1, region FROM cellsiterouter "
            "WHERE LOWER(hostname) = %s OR LOWER(loopback_0) = %s OR LOWER(loopback_1) = %s",
            (query_value, query_value, query_value),
        )
        csr_rows = cursor.fetchall()

        cursor.execute(
            "SELECT a.equipo_agregador, a.ip_gestion, a.ip_sistema, a.region, ipv.loopback0_ipv6 "
            "FROM agregattion a "
            "LEFT JOIN Agg_ipv6 ipv ON LOWER(ipv.ipran) = LOWER(a.equipo_agregador) "
            "WHERE LOWER(a.equipo_agregador) = %s OR LOWER(a.ip_sistema) = %s OR LOWER(a.ip_gestion) = %s",
            (query_value, query_value, query_value),
        )
        agg_rows = cursor.fetchall()

        results = [normalize_router(row, "cellsiterouter") for row in csr_rows]
        results += [normalize_router(row, "agregattion") for row in agg_rows]
        return results
    except Error:
        return None
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def find_router(router_name):
    router_name = (router_name or "").strip()
    if not router_name:
        return None, "Debe escribir el nombre, Loopback0, Loopback1 o región."

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        query_value = router_name.lower()

        cursor.execute(
            "SELECT hostname, loopback_0, loopback_1, region FROM cellsiterouter "
            "WHERE LOWER(hostname) = %s OR LOWER(loopback_0) = %s OR LOWER(loopback_1) = %s",
            (query_value, query_value, query_value),
        )
        csr_rows = cursor.fetchall()

        cursor.execute(
            "SELECT a.equipo_agregador, a.ip_gestion, a.ip_sistema, a.region, ipv.loopback0_ipv6 "
            "FROM agregattion a "
            "LEFT JOIN Agg_ipv6 ipv ON LOWER(ipv.ipran) = LOWER(a.equipo_agregador) "
            "WHERE LOWER(a.equipo_agregador) = %s OR LOWER(a.ip_sistema) = %s OR LOWER(a.ip_gestion) = %s",
            (query_value, query_value, query_value),
        )
        agg_rows = cursor.fetchall()

        results = [normalize_router(row, "cellsiterouter") for row in csr_rows]
        results += [normalize_router(row, "agregattion") for row in agg_rows]

        if results:
            if len(results) == 1:
                return {"type": "single", "item": results[0], "title": "Router encontrado"}, None
            return {"type": "list", "rows": results, "title": "Resultados encontrados"}, None

        cursor.execute(
            "SELECT hostname, loopback_0, loopback_1, region FROM cellsiterouter "
            "WHERE LOWER(region) = %s",
            (query_value,),
        )
        csr_rows = cursor.fetchall()

        cursor.execute(
            "SELECT a.equipo_agregador, a.ip_gestion, a.ip_sistema, a.region, ipv.loopback0_ipv6 "
            "FROM agregattion a "
            "LEFT JOIN Agg_ipv6 ipv ON LOWER(ipv.ipran) = LOWER(a.equipo_agregador) "
            "WHERE LOWER(a.region) = %s",
            (query_value,),
        )
        agg_rows = cursor.fetchall()

        results = [normalize_router(row, "cellsiterouter") for row in csr_rows]
        results += [normalize_router(row, "agregattion") for row in agg_rows]

        if results:
            return {
                "type": "list",
                "rows": results,
                "title": f"Routers en región '{router_name}'",
            }, None

        return None, f"No se encontró ningún router para '{router_name}'."

    except Error as exc:
        return None, f"Error de conexión a la base de datos: {exc}"

    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


def find_multiple_routers(router_input):
    """Busca múltiples routers a partir de una entrada con nombres separados por línea o coma."""
    if not router_input:
        return None, "Debe escribir al menos un nombre de router."
    
    router_names = re.split(r'[,\n]', router_input)
    router_names = [name.strip() for name in router_names if name.strip()]
    
    if not router_names:
        return None, "Debe escribir al menos un nombre de router."
    
    all_results = []
    for name in router_names:
        results = find_single_router(name)
        if results:
            all_results.extend(results)
    
    if all_results:
        return {
            "type": "list",
            "rows": all_results,
            "title": f"Búsqueda múltiple: {len(router_names)} router(s) consultados, {len(all_results)} resultado(s) encontrados",
        }, None
    
    return None, f"No se encontraron routers para los términos especificados."


@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    error = None
    agg_script = None
    agg_error = None
    rcsr_script = None
    rcsr_error = None
    data_error = None
    preagg_script = None
    preagg_error = None
    router_name = ""
    data_form = {
        "topology": "1",
        "agg_hostname": "",
        "agg_port": "",
        "agg_loopback0": "",
        "agg_loopback0_ipv6": "",
        "agg_loopback1": "",
        "agg_region": "",
        "agg_zone": "",
        "agg_community": "",
        "preagg_hostname": "",
        "preagg_port": "",
        "preagg_loopback0": "",
        "csr_hostname": "",
        "csr_port": "",
        "rcsr_hostname": "",
        "rcsr_port": "",
        "wan_ip": "",
        "wan_ipv6": "",
        "csr_loopback0": "",
        "csr_loopback0_ipv6": "",
        "csr_loopback1": "",
        "locator": "",
        "ospf_area": "",
        "cli": "",
        "cidpext": "",
        "ide": "",
        "bulk_input": "",
        "mpls_password": "B&k9DO%d4",
    }

    active_section = "section-search"
    if request.method == "POST":
        mode = request.form.get("mode", "search")
        if mode == "data":
            active_section = "section-data"
            action = request.form.get("action", "generate")
            for key in data_form:
                data_form[key] = request.form.get(key, "") or data_form[key]

            if action == "fetch_agg":
                agg_info, data_error = get_agregator_info(data_form.get("agg_hostname"))
                if agg_info:
                    data_form.update(agg_info)
            elif action == "fetch_preagg":
                preagg_info, data_error = get_preagg_info(data_form.get("preagg_hostname"))
                if preagg_info:
                    data_form.update(preagg_info)
            elif action == "load_bulk":
                topology = data_form.get("topology", "1")
                bulk_data = parse_bulk_input(data_form.get("bulk_input", ""))
                if topology == "2":
                    agg_host = bulk_data.get("agg_hostname", "")
                    is_preagg_row = bool(agg_host and agg_host.upper().startswith("RCSR"))
                    if is_preagg_row:
                        if bulk_data.get("agg_hostname"):
                            data_form["preagg_hostname"] = bulk_data["agg_hostname"]
                            data_form["agg_hostname"] = ""
                        if bulk_data.get("agg_port"):
                            data_form["preagg_port"] = bulk_data["agg_port"]
                            data_form["agg_port"] = ""
                        if bulk_data.get("agg_loopback0"):
                            data_form["preagg_loopback0"] = bulk_data["agg_loopback0"]
                            data_form["agg_loopback0"] = ""
                    else:
                        data_form["preagg_hostname"] = ""
                        data_form["preagg_port"] = ""
                        data_form["preagg_loopback0"] = ""
                        if bulk_data.get("agg_hostname"):
                            data_form["agg_hostname"] = bulk_data["agg_hostname"]
                            # Auto-complete AGG data in topology 2
                            agg_info, _ = get_agregator_info(bulk_data["agg_hostname"])
                            if agg_info:
                                data_form.update(agg_info)
                        if bulk_data.get("agg_port"):
                            data_form["agg_port"] = bulk_data["agg_port"]
                        if bulk_data.get("agg_loopback0"):
                            data_form["agg_loopback0"] = bulk_data["agg_loopback0"]
                else:  # topology == "1"
                    if bulk_data.get("agg_hostname"):
                        data_form["agg_hostname"] = bulk_data["agg_hostname"]
                        # Auto-complete AGG data in topology 1
                        agg_info, _ = get_agregator_info(bulk_data["agg_hostname"])
                        if agg_info:
                            data_form.update(agg_info)
                    if bulk_data.get("agg_port"):
                        data_form["agg_port"] = bulk_data["agg_port"]
                    if bulk_data.get("agg_loopback0"):
                        data_form["agg_loopback0"] = bulk_data["agg_loopback0"]
                for key, value in bulk_data.items():
                    if value and key not in {"agg_hostname", "agg_port", "agg_loopback0"}:
                        data_form[key] = value
                if bulk_data.get("has_ipv6") and not data_form.get("agg_loopback0_ipv6"):
                    data_form["agg_loopback0_ipv6"] = "---"
            elif action == "generate":
                topology = data_form.get("topology", "1")
                if topology == "1":
                    agg_script, agg_error = generate_agg_script(data_form)
                    rcsr_script, rcsr_error = generate_rcsr_script(data_form)
                    if agg_error or rcsr_error:
                        data_error = " ".join([e for e in [agg_error, rcsr_error] if e])
                    else:
                        active_section = "section-script-agg"
                elif topology == "2":
                    agg_script, agg_error = generate_agg_with_rcsr_peer(data_form)
                    preagg_script, preagg_error = generate_preagg_script(data_form)
                    rcsr_script, rcsr_error = generate_rcsr_script_agg_bgp(data_form)
                    if agg_error or preagg_error or rcsr_error:
                        data_error = " ".join([e for e in [agg_error, preagg_error, rcsr_error] if e])
                    else:
                        active_section = "section-script-agg"
            elif action == "generate_excel":
                topology = data_form.get("topology", "1")
                rcsr_name = (data_form.get("rcsr_hostname") or data_form.get("csr_hostname") or "script").strip()
                rcsr_name = re.sub(r'[^A-Za-z0-9_-]+', '_', rcsr_name)
                download_name = f"{rcsr_name}_script.xlsx"

                if topology == "1":
                    agg_script, _ = generate_agg_script(data_form)
                    rcsr_script, _ = generate_rcsr_script(data_form)
                    excel_data = create_excel([("AGG", agg_script), ("rCSR", rcsr_script)])
                elif topology == "2":
                    agg_script, _ = generate_agg_with_rcsr_peer(data_form)
                    preagg_script, _ = generate_preagg_script(data_form)
                    rcsr_script, _ = generate_rcsr_script_agg_bgp(data_form)
                    excel_data = create_excel(
                        [("AGG", agg_script), ("Preagregador", preagg_script), ("rCSR", rcsr_script)],
                    )
                else:
                    agg_script, _ = generate_agg_script(data_form)
                    rcsr_script, _ = generate_rcsr_script(data_form)
                    excel_data = create_excel([("AGG", agg_script), ("rCSR", rcsr_script)])

                return send_file(
                    excel_data,
                    as_attachment=True,
                    download_name=download_name,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        else:
            router_name = request.form.get("router_name", "")
            result, error = find_multiple_routers(router_name)

    return render_template(
        "index.html",
        result=result,
        error=error,
        router_name=router_name,
        data_form=data_form,
        agg_script=agg_script,
        agg_error=agg_error,
        rcsr_script=rcsr_script,
        rcsr_error=rcsr_error,
        data_error=data_error,
        preagg_script=preagg_script if 'preagg_script' in locals() else None,
        preagg_error=preagg_error if 'preagg_error' in locals() else None,
        active_section=active_section,
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5051)
